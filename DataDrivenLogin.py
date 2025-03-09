from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

# Load the Excel Sheet

workbook = load_workbook('testdata.xlsx')

# Selecting Active sheet
sheet = workbook.active

driver = webdriver.Chrome()

driver.maximize_window()

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]
    
    driver.get('https://www.saucedemo.com/')
    time.sleep(5)
    driver.find_element(By.ID, value='user-name').send_keys(username)
    driver.find_element(By.ID, value='password').send_keys(password)
    driver.find_element(By.ID, value='login-button').click()
    time.sleep(5)
    driver.find_element(By.XPATH, value='//*[@id="item_4_title_link"]/div').click()
    time.sleep(5)
    driver.find_element(By.XPATH, value='//*[@id="logout_sidebar_link"]')
    time.sleep(5)

driver.quit()
